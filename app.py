from flask import Flask, render_template, request, redirect, url_for, session
import os
import base64
import datetime
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook, Workbook
from pydub import AudioSegment  # Import pydub for audio processing
import traceback

app = Flask(__name__)

# Configuration
PROJECT_FOLDER = "./"
USER_DATA_FILE = f"{PROJECT_FOLDER}/User_Data.xlsx"
USER_RECORDS_FOLDER = f"{PROJECT_FOLDER}/User_Records/"
GENERATED_AUDIO_FOLDER = f"{PROJECT_FOLDER}/Generated_Audio/"
TEXT_INPUT_FILE = f"{PROJECT_FOLDER}/Text_Input.xlsx"
API_KEY = "W1vp8RVy2tnAw0GEj0NPqRszlWIXCfiDyLR5qOsY1rw="  # Your Speechify API key

# Create directories if they don't exist
os.makedirs(USER_RECORDS_FOLDER, exist_ok=True)
os.makedirs(GENERATED_AUDIO_FOLDER, exist_ok=True)

# Initialize Excel files
if not os.path.exists(USER_DATA_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["User_ID", "Voice_ID", "Timestamp"])
    wb.save(USER_DATA_FILE)

if not os.path.exists(TEXT_INPUT_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Text", "File_name"])
    ws.append(["This is a test sentence.", "T001"])
    ws.append(["Another test sentence for audio.", "T002"])
    wb.save(TEXT_INPUT_FILE)

# Configure secret key for session
app.secret_key = "your_secret_key"

# Helper Functions
def get_emotion_choices():
    return {
        "None": None, "angry": "angry", "cheerful": "cheerful", "sad": "sad", "terrified": "terrified",
        "relaxed": "relaxed", "fearful": "fearful", "surprised": "surprised", "calm": "calm",
        "assertive": "assertive", "energetic": "energetic", "warm": "warm", "direct": "direct", "bright": "bright"
    }

def get_voice_id(name, audio_file):
    try:
        # Validate the audio file exists and is accessible
        if not os.path.exists(audio_file):
            print(f"Error in get_voice_id for {name}: Audio file {audio_file} does not exist")
            return None
        if os.path.getsize(audio_file) > 5 * 1024 * 1024:  # 5MB limit as per Speechify docs
            print(f"Error in get_voice_id for {name}: Audio file {audio_file} exceeds 5MB limit")
            return None

        # Speechify API endpoint for voice cloning
        url = "https://api.sws.speechify.com/v1/voices"
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Accept": "application/json"
        }

        # Set up a session with retry logic for transient network issues
        session = requests.Session()
        retries = Retry(
            total=3,  # Retry 3 times
            backoff_factor=1,  # Wait 1, 2, 4 seconds between retries
            status_forcelist=[429, 500, 502, 503, 504],  # Retry on these HTTP status codes
            allowed_methods=["POST"]
        )
        session.mount("https://", HTTPAdapter(max_retries=retries))

        # Prepare the audio file and metadata for the API request
        with open(audio_file, "rb") as f:
            files = {
                "sample": f  # Correct field name as per Speechify API
            }
            data = {
                "name": name,
                "consent": '{"fullName": "User", "email": "user@example.com"}'  # Correct consent format
            }

            print(f"Sending voice cloning request to Speechify for {name} at {url}...")
            # Send the request with a timeout
            response = session.post(url, headers=headers, files=files, data=data, timeout=30)

        # Check if the request was successful
        if response.status_code != 200:
            print(f"Error in get_voice_id for {name}: HTTP {response.status_code} - {response.text}")
            return None

        # Parse the JSON response
        response_data = response.json()
        voice_id = response_data.get("id")  # Correct field name as per Speechify API
        if not voice_id:
            print(f"Error in get_voice_id for {name}: No id in response - {response_data}")
            return None

        print(f"Received voice ID for {name}: {voice_id}")
        return voice_id

    except requests.exceptions.NameResolutionError as e:
        print(f"Error in get_voice_id for {name}: DNS resolution failed - {str(e)}")
        print("Possible causes: Incorrect hostname, no internet access, or DNS server issues.")
        print("Action: Check your network connection, verify the API endpoint, or try again later.")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"Error in get_voice_id for {name}: Connection error - {str(e)}")
        print("Possible causes: Network outage, firewall blocking, or API server down.")
        print("Action: Check your network connection or try again later.")
        return None
    except requests.exceptions.Timeout as e:
        print(f"Error in get_voice_id for {name}: Request timed out - {str(e)}")
        print("Action: Check the API server status or increase the timeout value.")
        return None
    except Exception as e:
        print(f"Error in get_voice_id for {name}: Unexpected error - {str(e)}")
        return None

def text_to_speech_speechify(text, voice_id, filename, emotion=None, rate="0%"):
    """
    Convert text to speech using Speechify API and adjust speed with pydub.
    
    Args:
        text (str): Text to convert to speech
        voice_id (str): Speechify voice ID
        filename (str): Full output audio file path
        emotion (str, optional): Emotion for voice synthesis
        rate (str, optional): Speech rate as a percentage string (e.g., "-20%", "30%")
    
    Returns:
        str or None: Path to generated audio file, or None if generation fails
    """
    try:
        # Validate inputs
        if not text:
            print("Error: No text provided for speech synthesis")
            return None
        
        if not voice_id:
            print("Error: No voice ID specified")
            return None
        
        # Normalize the file path for consistency across operating systems
        filename = os.path.normpath(filename)
        
        # Ensure the directory exists
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        # Prepare request data
        headers = {
            "Authorization": f"Bearer {API_KEY}", 
            "Content-Type": "application/json"
        }
        
        # Validate the rate format (e.g., "-20%", "30%")
        try:
            # Ensure the rate ends with "%"
            if not rate.endswith("%"):
                rate = f"{rate}%"
            # Extract the numerical part and validate the range
            rate_value = int(rate.replace("%", ""))
            if not -50 <= rate_value <= 50:
                print(f"Error: Speech rate {rate_value}% is out of range (-50% to +50%)")
                return None
        except ValueError:
            print(f"Error: Invalid speech rate format: {rate}")
            return None
        
        # Convert the percentage to a speed multiplier
        # e.g., -20% -> 0.8 (80% of normal speed), +30% -> 1.3 (130% of normal speed)
        speed_multiplier = 1.0 + (rate_value / 100.0)
        
        # Construct data payload (without speech_rate, since we'll adjust speed locally)
        data = {
            "input": text,
            "voice_id": voice_id,
            "audio_format": "mp3"
        }
        
        # Add emotion if provided
        if emotion and emotion.lower() != "none":
            data["emotion"] = emotion
        
        # Log the payload for debugging
        print(f"Sending API request with payload: {data}")
        
        # Make API request
        response = requests.post(
            "https://api.sws.speechify.com/v1/audio/speech", 
            headers=headers, 
            json=data
        )
        
        # Check response
        if response.status_code == 200:
            # Ensure filename has .mp3 extension
            if not filename.lower().endswith('.mp3'):
                filename = f"{filename}.mp3"
            
            # Decode and save the audio file
            audio_data = response.json().get("audio_data")
            if not audio_data:
                print("Error: No audio data received from Speechify API")
                return None
            
            # Save the original audio file temporarily
            temp_filename = filename + ".temp.mp3"
            with open(temp_filename, "wb") as f:
                f.write(base64.b64decode(audio_data))
            
            # Verify file was created
            if not (os.path.exists(temp_filename) and os.path.getsize(temp_filename) > 0):
                print("Error: Failed to create a valid temporary audio file")
                return None
            
            # Load the audio file with pydub
            audio = AudioSegment.from_mp3(temp_filename)
            
            # Adjust the playback speed
            if speed_multiplier != 1.0:
                # Export to WAV for processing
                wav_temp = filename + ".temp.wav"
                audio.export(wav_temp, format="wav")
                
                # Load the WAV file and adjust speed
                audio = AudioSegment.from_wav(wav_temp)
                new_frame_rate = int(audio.frame_rate * speed_multiplier)
                audio = audio._spawn(audio.raw_data, overrides={"frame_rate": new_frame_rate})
                
                # Export the adjusted audio back to MP3
                audio.export(filename, format="mp3")
                
                # Clean up temporary files
                os.remove(wav_temp)
            else:
                # If speed_multiplier is 1.0, just rename the temp file
                os.rename(temp_filename, filename)
            
            # Clean up the temporary MP3 file if it still exists
            if os.path.exists(temp_filename):
                os.remove(temp_filename)
            
            # Verify the final file was created
            if os.path.exists(filename) and os.path.getsize(filename) > 0:
                print(f"Successfully generated audio file with adjusted speed: {filename}")
                return filename
            else:
                print("Error: Failed to create a valid audio file after speed adjustment")
                return None
        
        else:
            print(f"API Error: Status code {response.status_code}")
            print(f"Response: {response.text}")
            return None
    
    except Exception as e:
        print(f"Unexpected error in text-to-speech conversion: {e}")
        import traceback
        traceback.print_exc()
        return None

def save_user_data(user_id, voice_id):
    try:
        wb = load_workbook(USER_DATA_FILE)
        ws = wb.active
        ws.append([user_id, voice_id, datetime.datetime.now().strftime("%Y-%m-d %H:%M:%S")])
        wb.save(USER_DATA_FILE)
        print(f"Saved user data: {user_id, voice_id}")
    except Exception as e:
        print(f"Error saving user data for {user_id}: {str(e)}")
        raise  # Re-raise the exception to catch it in the calling function

def load_text_inputs():
    try:
        wb = load_workbook(TEXT_INPUT_FILE)
        ws = wb.active
        texts = {}
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if len(row) >= 2 and row[0] and row[1]:  # Ensure both text and file_name are non-empty
                text = str(row[0]).strip()
                file_name = str(row[1]).strip()
                text_id = f"T{idx:03d}"
                texts[text_id] = {"text": text, "file_name": file_name}
                print(f"Loaded text input: {text_id, text, file_name}")
        return texts
    except Exception as e:
        print(f"Error loading text inputs: {str(e)}")
        return {}

# Routes
@app.route('/', methods=['GET', 'POST'])
def index():
    message = ""
    print("Received request:", request.method)
    if request.method == 'POST':
        print("Form data:", request.form)
        print("Files:", request.files)

        # Initialize session variable for uploaded files if not already set
        if 'uploaded_files' not in session:
            session['uploaded_files'] = []

        if 'upload' in request.form:
            print("Upload button clicked")
            audio_files = request.files.getlist('audio_files')
            print("Uploading files:", [f.filename for f in audio_files])
            if not audio_files or all(not f.filename for f in audio_files):
                message = "No files selected. Please choose at least one MP3 file."
                print(message)
            else:
                for audio_file in audio_files:
                    if audio_file and audio_file.filename:
                        try:
                            filepath = os.path.join(USER_RECORDS_FOLDER, audio_file.filename)
                            audio_file.save(filepath)
                            session['uploaded_files'].append(filepath)  # Add to session
                            print(f"Saved audio file: {filepath}")
                        except Exception as e:
                            print(f"Error saving file {audio_file.filename}: {str(e)}")
                            message += f"Failed to save {audio_file.filename}: {str(e)}<br>"
                if not message:
                    message = "Audio files uploaded successfully!"
                    print(message)

        elif 'generate' in request.form:
            print("Generate button clicked")
            emotion = request.form['emotion']
            # The rate from the form is a numerical value (e.g., "-20", "30")
            rate = request.form.get('rate', '0')  # Default to '0' if not provided

            # Validate and format the rate
            try:
                rate_value = int(rate)  # Convert to integer to validate
                if not -50 <= rate_value <= 50:
                    message = f"Error: Speech rate {rate_value}% is out of range (-50% to +50%)."
                    print(message)
                else:
                    rate = f"{rate_value}%"  # Convert to string with "%" (e.g., "-20%")
                    print(f"Selected emotion: {emotion}, rate: {rate}")
            except ValueError:
                message = f"Error: Invalid speech rate format: {rate}. Please select a value between -50 and 50."
                print(message)
                rate = "0%"  # Fallback to default rate

            user_voice_ids = {}

            # Process only the files uploaded in this session
            uploaded_files = session.get('uploaded_files', [])
            if not uploaded_files:
                message = "No files uploaded in this session. Please upload files first."
                print(message)
            else:
                for filepath in uploaded_files:
                    filename = os.path.basename(filepath)
                    user_id = filename.rsplit(".", 1)[0]  # Split on the last "." to handle filenames with multiple dots
                    print(f"Processing file: {filename}, user_id: {user_id}, filepath: {filepath}")
                    try:
                        voice_id = get_voice_id(user_id, filepath)
                        print(f"Voice ID for {user_id}: {voice_id}")
                        if voice_id:
                            user_voice_ids[user_id] = voice_id
                            save_user_data(user_id, voice_id)
                        else:
                            message += f"Failed to get voice ID for {user_id}<br>"
                            print(f"Failed to get voice ID for {user_id}: voice_id is None")
                    except Exception as e:
                        print(f"Error processing voice ID for {user_id}: {str(e)}")
                        message += f"Failed to process voice ID for {user_id}: {str(e)}<br>"

                # Generate Audio
                print("Loading text inputs...")
                texts = load_text_inputs()
                if not texts:
                    message = "No text inputs found in Text_Input.xlsx. Please add text data."
                    print(message)
                else:
                    if not user_voice_ids:
                        message = "No voice IDs generated. Cannot generate audio."
                        print(message)
                    else:
                        for user_id, voice_id in user_voice_ids.items():
                            user_folder = os.path.join(GENERATED_AUDIO_FOLDER, user_id)
                            os.makedirs(user_folder, exist_ok=True)
                            print(f"Created user folder: {user_folder}")
                            for text_id, data in texts.items():
                                # Modify the file name to include emotion and rate
                                base_file_name = data['file_name']  # e.g., "T001"
                                # Handle emotion in the file name
                                emotion_part = emotion.lower() if emotion and emotion.lower() != "none" else "neutral"
                                # Handle rate in the file name (e.g., "-20%" -> "minus20", "30%" -> "plus30")
                                rate_part = f"{'minus' if rate_value < 0 else 'plus'}{abs(rate_value)}" if rate_value != 0 else "normal"
                                # Construct the new file name (e.g., "T001_cheerful_minus20.mp3")
                                new_file_name = f"{base_file_name}_{emotion_part}_{rate_part}.mp3"
                                output_path = os.path.join(user_folder, new_file_name)
                                
                                print(f"Generating audio for {user_id}, text: {data['text']}, emotion: {emotion}, rate: {rate}, output: {new_file_name}")
                                filepath = text_to_speech_speechify(
                                    data["text"],
                                    voice_id,
                                    output_path,
                                    emotion,
                                    rate
                                )
                                if filepath:
                                    message += f"Generated {new_file_name} for {user_id}<br>"
                                else:
                                    message += f"Failed to generate {new_file_name} for {user_id}<br>"

            # Clear uploaded files from session after processing
            session.pop('uploaded_files', None)

    emotions = get_emotion_choices()
    return render_template('index.html', emotions=emotions, message=message)

if __name__ == '__main__':
    app.run(debug=True, port=5000)